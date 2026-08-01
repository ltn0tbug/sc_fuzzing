"""Build a combined results_summary.xlsx spanning BOTH datasets.

Mirrors the original SmartBugs-only `results_summary.xlsx` layout (an `_overview`
sheet + one per-method sheet), but:
  * adds a `dataset` column so SmartBugs + DeFiHackLabs rows coexist, and
  * the `_overview` sheet carries one row per (dataset, method).

Per-contract rich columns (branch/line/function coverage, max_reward, tokens)
are reconstructed from the full run-log JSONs; status/elapsed/category/target
come from each cell's `_summary.json`.

    uv run python utils/build_summary_xlsx.py
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

RESULTS = Path(__file__).resolve().parents[1] / "output" / "experiment"
OUT = RESULTS / "results_summary_combined.xlsx"

METHODS = ["sscfuzz", "rlfuzz", "madfuzz", "randomfuzz", "llmfuzz"]
# (dataset label, results subdir relative to RESULTS — one per dataset)
DATASETS = [("smartbugs", "smartbugs"), ("defihacklabs", "defihacklabs")]

COLS = [
    "dataset", "contract_id", "category", "target", "status", "fail_reason",
    "elapsed_s",
    "bc_branches_hit", "bc_branches_total", "bc_branch_coverage",
    "branches_hit", "branches_total", "branch_coverage",
    "lines_hit", "lines_total", "line_coverage",
    "functions_hit", "functions_total", "function_coverage",
    "max_reward", "total_reward", "bugs_found",
    "input_tokens", "output_tokens", "total_tokens",
]


def _ratio(hit: int, total: int) -> float:
    return round(hit / total, 4) if total else 0.0


def _enrich_from_runlog(path: Path) -> dict:
    """Pull branch/line/function coverage + tokens + max_reward from a run-log."""
    out = {k: None for k in (
        "branches_hit", "branches_total", "branch_coverage",
        "lines_hit", "lines_total", "line_coverage",
        "functions_hit", "functions_total", "function_coverage",
        "bc_branches_hit", "bc_branches_total", "bc_branch_coverage",
        "max_reward", "total_reward", "bugs_found",
        "input_tokens", "output_tokens", "total_tokens",
    )}
    if not path.exists():
        return out
    d = json.loads(path.read_text())
    s = d.get("summary", {})

    out["bc_branches_hit"] = s.get("total_coverage_bc_branches")
    out["bc_branches_total"] = s.get("total_bc_branches")
    out["bc_branch_coverage"] = s.get("bc_coverage_ratio")
    out["branches_hit"] = s.get("total_coverage_branches")
    out["branches_total"] = s.get("total_branches")
    out["branch_coverage"] = s.get("coverage_ratio")
    out["total_reward"] = s.get("total_reward")
    out["bugs_found"] = s.get("total_bugs_found")

    tok = s.get("token_usage", {}) or {}
    out["input_tokens"] = tok.get("input_tokens", 0)
    out["output_tokens"] = tok.get("output_tokens", 0)
    out["total_tokens"] = tok.get("total_tokens", 0)

    # Cumulative line / function coverage: union the per-run sets; the run-log
    # has no cumulative line/function set, so we rebuild it from `*_this_run`.
    lines_hit: set = set()
    funcs_hit: set = set()
    lines_total = funcs_total = 0
    max_reward = None
    for it in d.get("iterations", []):
        fo = it.get("fuzzing_output", {}) or {}
        lines_hit |= set(fo.get("lines_this_run", []) or [])
        funcs_hit |= set(fo.get("functions_this_run", []) or [])
        if fo.get("lines_total"):
            lines_total = max(lines_total, fo["lines_total"])
        if fo.get("functions_total"):
            funcs_total = max(funcs_total, fo["functions_total"])
        r = fo.get("reward")
        if r is not None:
            max_reward = r if max_reward is None else max(max_reward, r)

    out["lines_hit"] = len(lines_hit)
    out["lines_total"] = lines_total
    out["line_coverage"] = _ratio(len(lines_hit), lines_total)
    out["functions_hit"] = len(funcs_hit)
    out["functions_total"] = funcs_total
    out["function_coverage"] = _ratio(len(funcs_hit), funcs_total)
    out["max_reward"] = round(max_reward, 4) if max_reward is not None else None
    return out


def _rows_for(method: str) -> list[dict]:
    rows: list[dict] = []
    for ds_label, subdir in DATASETS:
        cell = RESULTS / subdir / method if subdir else RESULTS / method
        summ_path = cell / "_summary.json"
        if not summ_path.exists():
            continue
        summ = json.loads(summ_path.read_text())
        for r in summ.get("results", []):
            cid = r["id"]
            log = cell / f"{cid.replace('/', '_')}.json"
            row = {
                "dataset": ds_label,
                "contract_id": cid,
                "category": r.get("category") or r.get("chain") or "",
                "target": r.get("target", ""),
                "status": "success" if r.get("status") == "ok" else r.get("status"),
                "fail_reason": r.get("fail_reason"),
                "elapsed_s": r.get("elapsed_s"),
            }
            enr = _enrich_from_runlog(log)
            # fall back to the compact summary numbers if the run-log is missing
            enr["bc_branches_hit"] = enr["bc_branches_hit"] if enr["bc_branches_hit"] is not None else r.get("bc_branches_hit")
            enr["bc_branches_total"] = enr["bc_branches_total"] if enr["bc_branches_total"] is not None else r.get("bc_branches_total")
            enr["bc_branch_coverage"] = enr["bc_branch_coverage"] if enr["bc_branch_coverage"] is not None else r.get("bc_coverage_ratio")
            enr["total_reward"] = enr["total_reward"] if enr["total_reward"] is not None else r.get("total_reward")
            enr["bugs_found"] = enr["bugs_found"] if enr["bugs_found"] is not None else r.get("bugs")
            row.update(enr)
            rows.append(row)
    return rows


HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_sheet(ws, header: list[str], rows: list[list]):
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    # autosize-ish
    for c in range(1, len(header) + 1):
        letter = get_column_letter(c)
        width = max(len(str(header[c - 1])),
                    *(len(str(r[c - 1])) for r in rows)) if rows else len(str(header[c - 1]))
        ws.column_dimensions[letter].width = min(max(width + 2, 8), 42)


def main():
    wb = openpyxl.Workbook()

    # ---- per-method sheets + collect for overview ----
    method_rows: dict[str, list[dict]] = {m: _rows_for(m) for m in METHODS}

    overview = wb.active
    overview.title = "_overview"
    ov_header = [
        "dataset", "method", "n_total", "n_success", "n_fail",
        "avg_bc_branch_cov", "avg_branch_cov", "avg_line_cov", "avg_function_cov",
        "total_bugs_found", "total_input_tokens", "total_output_tokens", "total_tokens",
    ]
    ov_rows = []
    for ds_label, _ in DATASETS:
        for m in METHODS:
            rs = [r for r in method_rows[m] if r["dataset"] == ds_label]
            if not rs:
                continue
            n = len(rs)
            n_succ = sum(1 for r in rs if r["status"] == "success")

            def _avg(key):
                vals = [r[key] for r in rs if isinstance(r.get(key), (int, float))]
                return round(sum(vals) / len(vals), 4) if vals else 0.0

            ov_rows.append([
                ds_label, m, n, n_succ, n - n_succ,
                _avg("bc_branch_coverage"), _avg("branch_coverage"),
                _avg("line_coverage"), _avg("function_coverage"),
                sum(r["bugs_found"] or 0 for r in rs),
                sum(r["input_tokens"] or 0 for r in rs),
                sum(r["output_tokens"] or 0 for r in rs),
                sum(r["total_tokens"] or 0 for r in rs),
            ])
    _write_sheet(overview, ov_header, ov_rows)

    for m in METHODS:
        ws = wb.create_sheet(m)
        rows = [[r.get(c) for c in COLS] for r in method_rows[m]]
        _write_sheet(ws, COLS, rows)

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Sheets: {wb.sheetnames}")
    for r in ov_rows:
        print("  ", r[:5], "bc=%.3f" % r[5], "bugs=%d" % r[9])


if __name__ == "__main__":
    main()
